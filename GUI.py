# -*- coding: utf-8 -*-
import glob
import tkinter as tk
from tkinter import ttk, filedialog
import tkinter.font as font
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import CellIsRule
import json
from datetime import date


class startWindow:
    """Primera finestra"""
    carpeta = ""
    excelcomptes = ""

    def __init__(self):
        """Inicialitza la primera pantalla"""

        # Declara la finestra de l'aplicació
        # Treu la barra de tk
        self.arrel = tk.Tk()
        self.arrel.overrideredirect(True)

        # Defineix dimensions de la finestra ample x alt 300x200
        # que se situarà en la coordenada x=500,y=50
        # Centrem la finestra a la pantalla
        amplada_finestra = 600
        altura_finestra = 200
        amplada_monitor = self.arrel.winfo_screenwidth()
        altura_monitor = self.arrel.winfo_screenheight()
        x = round(amplada_monitor / 2 - amplada_finestra / 2)
        y = round(altura_monitor / 2 - altura_finestra / 2)

        self.arrel.geometry(f'{amplada_finestra}x{altura_finestra}+{x}+{y}')

        # Frames
        # Crea un frame per a la barra nova de títol
        back_ground = '#1d1d1d'
        title_barframe = tk.Frame(self.arrel, width=535, height=20, bg=back_ground, relief='raised', bd=1, pady=3,
                                  highlightcolor=back_ground, highlightthickness=0)
        # crear frame per al botó tancar
        close_frame = tk.Frame(self.arrel, bg=back_ground, width=10, height=10, relief='raised', bd=1,
                               highlightcolor=back_ground, highlightthickness=0)

        # Crea un frame per a la sel·leció de la carpeta
        folder_frame = tk.Frame(self.arrel, bg=back_ground, width=555, height=200)

        # Configurar grid
        self.arrel.columnconfigure(0, weight=1)

        # Grid Frames
        title_barframe.grid(row=0, sticky=tk.EW)
        close_frame.grid(row=0, sticky=tk.NE)
        folder_frame.grid(row=1, sticky=tk.NSEW)

        # Widggets
        # Títol finestra
        title_name = tk.Label(title_barframe, text="Financial Control", bg=back_ground, fg='white')
        # Crea un botó per tancar a la barra de títol
        close_button = tk.Button(close_frame, text='x', command=self.arrel.destroy, bg=back_ground,
                                 activebackground="red", bd=0, font="bold", fg='white', activeforeground="white",
                                 highlightthickness=0)
        # Entrada any
        self.any = tk.StringVar()
        self.any.set(f'{date.today().year}')
        self.entry_any = ttk.Entry(folder_frame, textvariable=self.any, justify=tk.LEFT, width=6,
                                 background=back_ground)

        # Etiqueta carpeta
        folder_label = tk.Label(folder_frame, text="Sel·lecciona la carpeta d'excels",
                                bg=back_ground, fg='white', padx=15, pady=30)
        # Entrada text carpeta
        self.carpeta = tk.StringVar()
        entry_folder = ttk.Entry(folder_frame, textvariable=self.carpeta, justify=tk.LEFT, width=50,
                                 background=back_ground)
        # Botó per buscar carpeta
        button_font = font.Font(family="Helvetica", size=8, weight="bold")
        search_button = tk.Button(folder_frame, text='Buscar carpeta', bg='#b5b5b5', activebackground="#ffffff", bd=0,
                                  font="bold", fg='black', activeforeground="black", command=self.get_folder_path)
        search_button['font'] = button_font
        # Botó continuar
        continue_button = tk.Button(folder_frame, text='Continuar', bg='#b5b5b5', activebackground="#ffffff", bd=0,
                                    font="bold", fg='black', activeforeground="black", command=self.next_window)
        continue_button['font'] = button_font

        # Grid widgets
        title_name.grid(row=0, column=0, columnspan=7, sticky=tk.NS)
        close_button.grid(sticky=tk.NE)
        self.entry_any.grid(row=0, column=0, sticky=tk.W, padx=15,pady=10)
        folder_label.grid(row=1, column=0, sticky=tk.W)
        entry_folder.grid(row=1, column=1, sticky=tk.W)
        search_button.grid(row=1, column=2, sticky=tk.E, padx=5)
        continue_button.grid(row=2, column=2, sticky=tk.SW, padx=5, pady=20)

        # Esdeveniment amb bind per poder moure la finestra

        def move_window(event):
            """Dotar de moviment a la finestra"""
            self.arrel.geometry('+{0}+{1}'.format(event.x_root, event.y_root))

        # Els botons canvien de color al passar per damunt
        def change_on_hovering(event):
            """El botó canvia de color al passar per sobre"""
            close_button.configure(bg='red')

        def change_search_on_hovering(event):
            """El botó canvia de color al passar per sobre"""
            search_button.configure(bg='black', fg='white')

        def change_continue_on_hovering(event):
            """El botó canvia de color al passar per sobre"""
            continue_button.configure(bg='black', fg='white')

        def return_to_normal_state(event):
            """El botó torna al seu estat inicial"""
            close_button.configure(bg=back_ground)

        def return_search_to_normal_state(event):
            """El botó torna al seu estat inicial"""
            search_button.configure(bg='#b5b5b5', fg='black')

        def return_continue_to_normal_state(event):
            """El botó torna al seu estat inicial"""
            continue_button.configure(bg='#b5b5b5', fg='black')

        title_barframe.bind('<B1-Motion>', move_window)
        close_button.bind('<Enter>', change_on_hovering)
        close_button.bind('<Leave>', return_to_normal_state)
        search_button.bind('<Enter>', change_search_on_hovering)
        search_button.bind('<Leave>', return_search_to_normal_state)
        continue_button.bind('<Enter>', change_continue_on_hovering)
        continue_button.bind('<Leave>', return_continue_to_normal_state)

        self.arrel.mainloop()

    def get_folder_path(self):
        """Obte la ruta de la carpeta i la llista d'excels"""
        self.folder_path = tk.filedialog.askdirectory(initialdir=r"C:\Users\ferra\OneDrive\Comptes\Excels_caixa",
                                                      title="Sel·lecciona una carpeta")
        self.carpeta.set(self.folder_path)

    def next_window(self):
        """Sel·lecciona l'excel del any o el crea, tanca la primera finestra i obre la següent"""
        any = self.entry_any.get()
        llista_comptes = glob.glob('C:/Users/ferra/OneDrive/Comptes/*.xlsx')
        for xls in llista_comptes:
            if any in xls:
                startWindow.excelcomptes = xls
            else:
                nou_excel = Workbook()
                nou_excel.save(filename=f'C:/Users/ferra/OneDrive/Comptes/Comptes_{any}.xlsx')
                startWindow.excelcomptes = f'C:/Users/ferra/OneDrive/Comptes/Comptes_{any}.xlsx'

        ## -----Implementar error de carpeta----------
        startWindow.carpeta = self.carpeta.get()
        self.arrel.destroy()
        secondWindow(startWindow)


class secondWindow(startWindow):
    """Segona finestra"""
    def safexit(self):
        with open('classificacio.json') as json_file:
            json_decoded = json.load(json_file)


        if self.entry_key.get().capitalize() in json_decoded["classificacio"]:
            json_decoded["classificacio"][self.entry_key.get().capitalize()].append(self.entry_valor.get().lower())
        else:
            json_decoded["classificacio"][self.entry_key.get().capitalize()] = [self.entry_valor.get().lower()]

        with open('classificacio.json', 'w') as json_file:
            json.dump(json_decoded, json_file, sort_keys=True, indent=4, ensure_ascii=False)

        updated_list = []
        with open('classificacio.json') as json_file:
            data = json.load(json_file)
            dic = data["classificacio"]
            for x in dic:
                updated_list.append(x)
        rows = len(self.rows)
        for i in range(0, rows):
            self.opt[i][3].set_menu(f'SELECTION - {i}', *updated_list)


        self.dicw.destroy()

    def safe(self):
        with open('classificacio.json') as json_file:
            json_decoded = json.load(json_file)

        if self.entry_key.get().capitalize() in json_decoded["classificacio"]:
            json_decoded["classificacio"][self.entry_key.get().capitalize()].append(self.entry_valor.get().lower())
        else:
            json_decoded["classificacio"][self.entry_key.get().capitalize()] = [self.entry_valor.get().lower()]

        with open('classificacio.json', 'w') as json_file:
            json.dump(json_decoded, json_file, sort_keys=True, indent=4, ensure_ascii=False)

        self.entry_key.delete(0, tk.END)
        self.entry_valor.delete(0, tk.END)

    def newcategory(self):
        """Crea un nou concepte o categoria al diccionari"""
        # Creem la segona pantalla
        self.dicw = tk.Tk()
        self.dicw.overrideredirect(True)

        # Configurar grid
        self.dicw.rowconfigure(0, weight=1)
        self.dicw.columnconfigure(0, weight=1)

        self.amp_dicw = 495
        self.alt_dicw = 245
        amplada_monitor = self.dicw.winfo_screenwidth()
        altura_monitor = self.dicw.winfo_screenheight()
        x = round(amplada_monitor / 2 - self.amp_dicw / 2)
        y = round((altura_monitor - 50) / 2 - self.alt_dicw / 2)

        self.dicw.geometry(f'{self.amp_dicw}x{self.alt_dicw}+{x}+{y}')

        # Frames
        # Creem un frame general
        self.frame_gen = tk.Frame(self.dicw, bg="gray", width=self.amp_dicw, height=self.alt_dicw)
        self.frame_gen.grid(sticky=tk.NSEW)
        # Crea un frame per a la barra nova de títol
        back_ground = '#1d1d1d'
        title_barframef = tk.Frame(self.frame_gen, width=self.amp_dicw, height=20, bg=back_ground,
                                  relief='raised', bd=1,
                                  pady=3, highlightcolor=back_ground, highlightthickness=0)
        # crear frame per al botó tancar
        close_framef = tk.Frame(self.frame_gen, bg=back_ground, width=10, height=10, relief='raised', bd=1,
                               highlightcolor=back_ground, highlightthickness=0)
        # Crea un frame diccionari
        dic_frame = tk.Frame(self.frame_gen, bg=back_ground)
        # Grid Frames
        title_barframef.grid(row=0, sticky=tk.EW)
        close_framef.grid(row=0, sticky=tk.NE)
        dic_frame.grid(row=1, sticky=tk.NSEW)
        # Widggets
        # Títol finestra
        title_namef = tk.Label(title_barframef, text="Financial Control", bg=back_ground, fg='white')
        # Crea un botó per tancar a la barra de títol
        close_buttonf = tk.Button(close_framef, text='x', command=self.dicw.destroy, bg=back_ground,
                                 activebackground="red", bd=0, font="bold", fg='white',
                                 activeforeground="white",
                                 highlightthickness=0)
        # Labels diccionari
        # Etiqueta key
        e_keydic = tk.Label(dic_frame, text="Categoria:", bg=back_ground, fg='white', padx=15, pady=30)
        # Entrada key
        self.key_dic = tk.StringVar()
        self.entry_key = ttk.Entry(dic_frame, textvariable=self.key_dic, justify=tk.LEFT, width=50,
                                 background=back_ground)
        # Etiqueta valor
        e_valor = tk.Label(dic_frame, text="Valor:", bg=back_ground, fg='white', padx=15, pady=30)
        # Entrada valor
        self.valor_dic = tk.StringVar()
        self.entry_valor = ttk.Entry(dic_frame, textvariable=self.valor_dic, justify=tk.LEFT, width=50,
                                 background=back_ground)
        # Botó tancar i guardar diccionari
        button_save = tk.Button(dic_frame, text="Guardar", bg=back_ground, fg='white', command=self.safe)
        button_saveexit = tk.Button(dic_frame, text="Guardar i tancar", bg=back_ground, fg='white', command=self.safexit)
        # Grid widgets
        title_namef.grid(row=0, column=0, sticky=tk.NSEW)
        close_buttonf.grid(sticky=tk.NE)
        e_keydic.grid(row=0, column=0, sticky=tk.W)
        e_valor.grid(row=1, column=0, sticky=tk.W)
        self.entry_key.grid(row=0, column=1, sticky=tk.W)
        self.entry_valor.grid(row=1, column=1, sticky=tk.W)
        button_save.grid(row=2, column=3, sticky=tk.EW, pady=2, padx=5)
        button_saveexit.grid(row=3, column=3, sticky=tk.EW, pady=2, padx=5)

        def move_window(event):
            """Dotar de moviment a la finestra"""
            self.dicw.geometry('+{0}+{1}'.format(event.x_root, event.y_root))

        def change_on_hovering(event):
            """El botó canvia de color al passar per sobre"""
            close_buttonf.configure(bg='red')

        def return_to_normal_state(event):
            """El botó torna al seu estat inicial"""
            close_buttonf.configure(bg=back_ground)

        title_barframef.bind('<B1-Motion>', move_window)
        close_buttonf.bind('<Enter>', change_on_hovering)
        close_buttonf.bind('<Leave>', return_to_normal_state)

    def select_item(self, selection):
        """Assigna la classificació a l'element"""
        rows = len(self.rows)
        zip_list = list(zip(self.rows, self.rows_taula))
        for i in range(0, rows):
            if selection == self.var[i][3].get():
                self.ws_act.cell(row=zip_list[i][0], column=5).value = selection
        self.ex_comptes.save(filename=startWindow.excelcomptes)
    def taula(self):
        """Crea la taula dels conceptes per classificar"""
        # creem un canvas per allotjar la scrollbar
        # Creem un canvas al frame canvas
        canvas = tk.Canvas(self.canvas_frame, bg='#1d1d1d')
        canvas.grid(row=0, column=0, sticky=tk.NSEW)
        # fiquem el scrolbar al frame canvas
        vsb = tk.Scrollbar(self.canvas_frame, orient="vertical", command=canvas.yview)
        vsb.grid(row=0, column=1, sticky=tk.NS)
        canvas.configure(yscrollcommand=vsb.set)
        # Crea un frame per la taula de conceptes
        self.con_frame = tk.Frame(canvas, bg='#1d1d1d')
        canvas.create_window((0, 0), window=self.con_frame, anchor=tk.NW, width=900)
        # Afegim els títols de la taula i les files corresponents a conceptes sense classificar
        cb = '#1d1d1d'
        font_titol = font.Font(family="Helvetica", size=10, weight="bold")
        rows = len(self.rows)

        self.rows_taula = []
        for i in range(1, rows + 1):
            self.rows_taula.append(i)
        columns = 4
        # Per poder crear variables de la taula
        self.labels = [[tk.Label() for j in range(columns)] for i in range(rows)]
        self.var = [[tk.StringVar(self.con_frame) for j in range(columns)] for i in range(rows)]
        self.opt = [[ttk.OptionMenu(self.con_frame, self.var[i][3], *self.llista_clas) for j in range(columns)] for i in
                    range(rows)]

        self.labels[0][0] = tk.Label(self.con_frame, text="CONCEPTE", font=font_titol, bg=cb, fg='white')
        self.labels[0][0].grid(row=0, column=0, ipadx=50, ipady=10)
        self.labels[0][1] = tk.Label(self.con_frame, text="DATA", font=font_titol, bg=cb, fg='white')
        self.labels[0][1].grid(row=0, column=1, ipadx=50, ipady=10)
        self.labels[0][2] = tk.Label(self.con_frame, text="IMPORT", font=font_titol, bg=cb, fg='white')
        self.labels[0][2].grid(row=0, column=2, ipadx=50, ipady=10)
        self.labels[0][3] = tk.Label(self.con_frame, text="CLASSIFICACIÓ", font=font_titol, bg=cb, fg='white')
        self.labels[0][3].grid(row=0, column=3, ipadx=50, ipady=10)

        for i in range(0, rows):
            # taula
            font_lab = font.Font(family="Helvetica", size=9)
            self.labels[i][0] = tk.Label(self.con_frame, text=str(self.ws_act.cell(row=self.rows[i], column=1).value),
                                         font=font_lab, bg=cb, fg='white')
            self.labels[i][0].grid(row=i, column=0, ipadx=70, ipady=10)
            self.labels[i][1] = tk.Label(self.con_frame, text=str(self.ws_act.cell(row=self.rows[i], column=2).value),
                                         font=font_lab, bg=cb, fg='white')
            self.labels[i][1].grid(row=i, column=1, ipadx=70, ipady=10)
            self.labels[i][2] = tk.Label(self.con_frame, text=str(self.ws_act.cell(row=self.rows[i], column=3).value),
                                         font=font_lab, bg=cb, fg='white')
            self.labels[i][2].grid(row=i, column=2, ipadx=70, ipady=10)
            self.var[i][3] = tk.StringVar()
            style = ttk.Style()
            style.configure('my.TMenubutton', font=('Helvetica', 9),  background=cb, foreground="white",
                            highlightthickness=0)
            self.opt[i][3] = ttk.OptionMenu(self.con_frame, self.var[i][3], f'SELECTION - {i}', *self.llista_clas,
                                            command=self.select_item, style='my.TMenubutton')
            self.opt[i][3].grid(row=i, column=3, sticky=tk.EW, ipadx=70, ipady=5)

        # Update buttons frames idle tasks to let tkinter calculate buttons sizes
        self.con_frame.update_idletasks()

        # Recalcul del canvas per que mostri totes les files de conceptes
        nfiles = 0
        if len(self.rows) >= 12:
            nfiles = 12
        else:
            nfiles = len(self.rows)
        columns_width = sum([self.labels[0][j].winfo_width() for j in range(0, 4)])
        rows_height = sum([self.labels[i][0].winfo_height() for i in range(0, nfiles)])
        extra_amplada = 0
        extra_altura = 0
        if columns_width < self.amplada_finestra:
            extra_amplada = self.amplada_finestra - columns_width - 15
            columns_width = columns_width + extra_amplada
        if rows_height < self.altura_finestra - 100:
            extra_altura = self.altura_finestra - rows_height - 185
            rows_height = rows_height + extra_altura
        self.canvas_frame.config(width=columns_width + vsb.winfo_width(), height=rows_height)
        # Set the canvas scrolling region
        canvas.config(scrollregion=canvas.bbox("all"))

    def check_concept(self, excel):
        """Classifica els conceptes segons el diccionari i si no estan classificats crea una interfície per
        classificarlos """
        self.ws_act = excel.active
        # iterem per cada una de les files de l'excel i assignem un concepte
        for j in range(2, self.ws_act.max_row + 1):
            concept = str(self.ws_act.cell(row=j, column=1).value).lower()
            # iterem pel diccionari de conceptes
            with open('classificacio.json') as json_file:
                data = json.load(json_file)
                dic = data["classificacio"]
                for elem in dic:
                    for y in data["classificacio"][elem]:
                        if y in concept:
                            self.ws_act.cell(row=j, column=5).value = elem
        # Llista de conceptes sense classificar
        self.rows = []
        for j in range(2, self.ws_act.max_row + 1):
            if self.ws_act.cell(row=j, column=5).value is None:
                self.rows.append(j)
        # Creem la llista elements diccionari json
        self.llista_clas = []
        with open('classificacio.json') as json_file:
            data = json.load(json_file)
            dic = data["classificacio"]
            for x in dic:
                self.llista_clas.append(x)
        if len(self.rows) > 0:
            # Creem la segona pantalla
            self.sw = tk.Tk()
            self.sw.overrideredirect(True)

            # Configurar grid
            self.sw.rowconfigure(0, weight=1)
            self.sw.columnconfigure(0, weight=1)

            self.amplada_finestra = 920
            self.altura_finestra = 700
            amplada_monitor = self.sw.winfo_screenwidth()
            altura_monitor = self.sw.winfo_screenheight()
            x = round(amplada_monitor / 2 - self.amplada_finestra / 2)
            y = round((altura_monitor - 50) / 2 - self.altura_finestra / 2)

            self.sw.geometry(f'{self.amplada_finestra}x{self.altura_finestra}+{x}+{y}')

            # Frames
            # Creem un frame general
            self.frame_main = tk.Frame(self.sw, bg="gray", width=self.amplada_finestra, height=self.altura_finestra-50)
            self.frame_main.grid(sticky=tk.NSEW)
            # Crea un frame per a la barra nova de títol
            back_ground = '#1d1d1d'
            title_barframe = tk.Frame(self.frame_main, width=self.amplada_finestra, height=20, bg=back_ground,
                                      relief='raised', bd=1,
                                      pady=3, highlightcolor=back_ground, highlightthickness=0)
            # crear frame per al botó tancar
            close_frame = tk.Frame(self.frame_main, bg=back_ground, width=10, height=10, relief='raised', bd=1,
                                   highlightcolor=back_ground, highlightthickness=0)
            # Crea un frame gestió excel
            gestio_frame = tk.Frame(self.frame_main, bg=back_ground)
            # Creem un frame per al canvas que allotjarà la taula
            self.canvas_frame = tk.Frame(self.frame_main, bg=back_ground)
            # Crear frame botons barra inferior
            self.fbuttons = tk.Frame(self.frame_main, bg=back_ground,width=self.amplada_finestra, height=20, pady=3)

            # Grid Frames
            title_barframe.grid(row=0, sticky=tk.EW)
            close_frame.grid(row=0, sticky=tk.NE)
            gestio_frame.grid(row=1, sticky=tk.EW)
            self.canvas_frame.grid(row=2, column=0, sticky=tk.EW)
            self.canvas_frame.rowconfigure(0, weight=1)
            self.canvas_frame.columnconfigure(0, weight=1)
            self.canvas_frame.grid_propagate(False)
            self.fbuttons.grid(row=3, sticky=tk.EW)
            # Widggets
            # Títol finestra
            title_name = tk.Label(title_barframe, text="Financial Control", bg=back_ground, fg='white')
            # Crea un botó per tancar a la barra de títol
            close_button = tk.Button(close_frame, text='x', command=self.sw.destroy, bg=back_ground,
                                     activebackground="red", bd=0, font="bold", fg='white',
                                     activeforeground="white",
                                     highlightthickness=0)
            # Labels gestió excel
            label_gestionant = tk.Label(gestio_frame, text=f'Gestionant despeses del mes', bg=back_ground,
                                        fg='white', font=font.Font(family="Helvetica", size=15, weight="bold"),
                                        padx=20, pady=15)
            label_mes = tk.Label(gestio_frame, text=self.nom_fulla,
                                 font=font.Font(family="Helvetica", size=25, weight="bold"),
                                 padx=20, pady=10, bg=back_ground, fg='white')
            # Barra inferior
            nouconcepte = tk.Button(self.fbuttons, text="Nova classe i/o concepte", bg=back_ground, fg='white', command=self.newcategory)
            tancar_finestra = tk.Button(self.fbuttons, text="Tanca", bg=back_ground, fg='white', command=self.sw.destroy)
            # Grid widgets
            title_name.grid(row=0, column=0, sticky=tk.NS)
            close_button.grid(sticky=tk.NE)
            label_gestionant.pack(fill=tk.X)
            label_mes.pack(fill=tk.X)
            nouconcepte.pack(side='left', pady=5, padx=5)
            tancar_finestra.pack(side='right', pady=5, padx=5)

            def move_window(event):
                """Dotar de moviment a la finestra"""
                self.sw.geometry('+{0}+{1}'.format(event.x_root, event.y_root))

            def change_on_hovering(event):
                """El botó canvia de color al passar per sobre"""
                close_button.configure(bg='red')

            def return_to_normal_state(event):
                """El botó torna al seu estat inicial"""
                close_button.configure(bg=back_ground)

            title_barframe.bind('<B1-Motion>', move_window)
            close_button.bind('<Enter>', change_on_hovering)
            close_button.bind('<Leave>', return_to_normal_state)

            # creació taula
            self.taula()

            self.sw.mainloop()
        else:
            ##---------- missatge de tot ok ---------
            print(f'tot fet a {self.ws_act}')

    def combiexcel(self, llista):
        """Afegeix els excels del banc a un de sol"""
        # Carreguem l'excel de comptes
        self.ex_comptes = load_workbook(startWindow.excelcomptes)
        for document in llista:
            # Carreguem l'excel del banc
            ex_caixa = load_workbook(document)
            sheet_caixa = ex_caixa['in']

            # Creació fulla segons el mes
            data_mes = str(sheet_caixa['B4'].value)[0:10].split('-')[1]
            with open('classificacio.json') as mes:
                data = json.load(mes)
                dic = data["mes"]
            self.nom_fulla = dic[data_mes]
            if self.nom_fulla not in self.ex_comptes.sheetnames:
                ws1 = self.ex_comptes.create_sheet(self.nom_fulla)
                ws1.title = self.nom_fulla
                ws2 = self.ex_comptes.active = self.ex_comptes[self.nom_fulla]

                # calculate total number of rows and
                # columns in source Excel file
                self.mr = sheet_caixa.max_row
                self.mc = sheet_caixa.max_column

                # copying the cell values from source
                # Excel file to destination Excel file
                for i in range(3, self.mr + 1):
                    for j in range(1, self.mc + 1):
                        if sheet_caixa.cell(row=i, column=j).value is not None:
                            # reading cell value from source excel file
                            c = sheet_caixa.cell(row=i, column=j)
                            # writing the read value to destination excel file
                            ws2.cell(row=i - 2, column=j).value = c.value

                for i in range(2, self.mr - 1):
                    # Format data
                    split_data = "/".join(list(reversed(str(ws2.cell(row=i, column=2).value)[0:10].split('-'))))
                    ws2.cell(row=i, column=2).value = split_data
                    # Format import
                    if ws2.cell(row=i, column=3).value is not None:
                        imports = ".".join((str(ws2.cell(row=i, column=3).value).split('.')))
                        ws2.cell(row=i, column=3).value = float(imports)
                    # Format saldo
                    if ws2.cell(row=i, column=4).value is not None:
                        saldo = ".".join("".join(str(ws2.cell(row=i, column=4).value)[0:-3].split('.')).split(','))
                        ws2.cell(row=i, column=4).value = float(saldo)

                # format titols columnes
                a1 = ws2['A1']
                b1 = ws2['B1']
                c1 = ws2['C1']
                d1 = ws2['D1']
                e1 = ws2['E1']
                e1.value = "Classificació"

                a1.font = Font(bold=True, size=15)
                b1.font = Font(bold=True, size=15)
                c1.font = Font(bold=True, size=15)
                d1.font = Font(bold=True, size=15)
                e1.font = Font(bold=True, size=15)
                #Funcio classificacio conceptes
                self.check_concept(self.ex_comptes)
                # filtres
                maxrow = ws2.max_row
                ws2.auto_filter.ref = f"A1:E{maxrow}"
                ws2.auto_filter.add_filter_column(5, ["Menja", "Compres", "Transport"])
                ws2.auto_filter.add_sort_condition(f"B2:B{maxrow}")

                # Taula resum
                # Crear la taula a partir del diccionari
                ws2['G2'] = "TAULA RESUM"
                ws2['G3'] = "Classificació"
                ws2['H3'] = "€"
                key_list = self.llista_clas
                num_files = len(key_list)
                ws2[f'G{num_files + 4}'] = "Estalvis"
                for i in range(4, num_files + 4):
                    ws2.cell(row=i, column=7).value = key_list[i - 4]
                    # El nom de l'operació excel ha de ser en ingles
                    ws2[f'H{i}'] = f'=SUMIF(E2:E{maxrow},G{i},C2:C{maxrow})'
                # amplada de columna
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 12
                ws2.column_dimensions['C'].width = 15
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions['E'].width = 20
                ws2.column_dimensions['G'].width = 15
                ws2.column_dimensions['H'].width = 15
                # Format condicional
                red_font = styles.Font(size=11, color='9c0006')
                redFill = styles.PatternFill(bgColor='ffc7ce', fill_type='solid')
                greenFill = styles.PatternFill(bgColor='c6efce', fill_type='solid')
                green_font = styles.Font(size=11, color='006100')
                ws2.conditional_formatting.add(f'H4:H{num_files + 4}', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True,
                                                                    fill=redFill, font=red_font))
                ws2.conditional_formatting.add(f'H4:H{num_files + 4}', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True,
                                                                    fill=greenFill, font=green_font))
                ws2.merge_cells('G2:H2')
                # Cel·la estalvis
                ws2[f'H{num_files + 4}'] = f'=D{maxrow}-D2'
                # Mateix format
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                for r in ws2[f'G2:H{num_files + 4}']:
                    for cell in r:
                        cell.border = thin_border
                for r in ws2[f'H4:H{num_files + 4}']:
                    for cell in r:
                        cell.number_format = '0.00€'
                for r in ws2[f'C2:C{num_files + 4}']:
                    for cell in r:
                        cell.number_format = '0.00€'
                for r in ws2[f'D2:D{num_files + 4}']:
                    for cell in r:
                        cell.number_format = '0.00€'
                for r in ws2['G3:H3']:
                    for cell in r:
                        cell.font = Font(bold=True, size=11)
                ws2['H3'].alignment = Alignment(horizontal="center")

            self.ex_comptes.save(filename=startWindow.excelcomptes)

    def __init__(self, finestra1):
        """Inicialitza la segona finestra"""
        llista_exel = glob.glob(finestra1.carpeta + '/*.xlsx')
        self.combiexcel(llista_exel)


def main():
    mi_app = startWindow()


if __name__ == "__main__":
    main()
