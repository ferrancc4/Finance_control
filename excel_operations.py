import glob
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font
import Diccionari


# Llistar els excels d'una carpeta
def listexcel():
    # Llistem els excels amb glob i creem un esdeveniment amb tkinter per sel·lecionar la carpeta

    window = tk.Tk()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    # calculate position x and y coordinates
    x = round((screen_width / 2) - (600 / 2))
    y = round((screen_height / 2) - (300 / 2))
    window.geometry(f'550x350+{x}+{y}')

    filepath = tk.filedialog.askdirectory(initialdir=r"C:\Users\ferra\OneDrive\Tesla\Economia", title="Select folder")
    window.destroy()
    window.mainloop()

    file_list = glob.glob(filepath + '/*.xlsx')
    return file_list


# Funcio iterar concepte per classificarlo

def check_concept(diccionari, excel):
    # Funcions per fer finestra per classificar conceptes

    def select_item(event):
        widget = event.widget
        selection = widget.curselection()
        valor_clas = widget.get(selection[0])
        ws_act.cell(row=j, column=5).value = valor_clas
        excel.save(filename='/EstatComptes.xlsx')

    # Classificació dels conseptes per cada una de les fulles
    ws_act = excel.active

    # iterem per cada una de les files
    for j in range(2, ws_act.max_row + 1):
        concept = str(ws_act.cell(row=j, column=1).value).lower()
        # iterem pel diccionari de conceptes
        for elem in diccionari.values():
            for res in elem:
                if res in concept:
                    ws_act.cell(row=j, column=5).value = list(diccionari.keys())[
                        list(diccionari.values()).index(elem)]
    # Llista de conceptes sense classificar
    for j in range(2, ws_act.max_row + 1):
        if ws_act.cell(row=j, column=5).value is None:
            window_grid = tk.Tk()
            window_grid.title('Classificació element')

            screen_width = window_grid.winfo_screenwidth()
            screen_height = window_grid.winfo_screenheight()
            # calculate position x and y coordinates
            x = round((screen_width / 2) - (600 / 2))
            y = round((screen_height / 2) - (300 / 2))
            window_grid.geometry(f'550x350+{x}+{y}')

            # Etiqueta Fulla
            fulla = str(ws_act['B4'].value)[3:5]
            nom_fulla = Diccionari.mes.get(fulla)
            fulla_label = tk.Label(window_grid, text=f'{nom_fulla}')
            fulla_label.grid(column=0, row=0, sticky=tk.W, padx=15, pady=5)

            # Etiqueta concepte
            concepte_label = tk.Label(window_grid, text='CONCEPTE')
            concepte_label.grid(column=0, row=1, sticky=tk.W, padx=15, pady=5)

            # Etiqueta import
            import_label = tk.Label(window_grid, text='IMPORT')
            import_label.grid(column=1, row=1, sticky=tk.W, padx=15, pady=5)

            # Etiqueta Data
            data_label = tk.Label(window_grid, text='Data')
            data_label.grid(column=2, row=1, sticky=tk.W, padx=15, pady=5)

            # Etiqueta classificació
            concepte_label = tk.Label(window_grid, text='CLASSIFICACIÓ')
            concepte_label.grid(column=3, row=1, sticky=tk.W, padx=15, pady=5)

            # Valor concepte
            vconcept = ws_act.cell(row=j, column=1).value
            concepte_label = tk.Label(window_grid, text=vconcept)
            concepte_label.grid(column=0, row=2, sticky=tk.E, padx=15, pady=5)

            # Valor import
            vconcept = ws_act.cell(row=j, column=3).value
            concepte_label = tk.Label(window_grid, text=vconcept)
            concepte_label.grid(column=1, row=2, sticky=tk.E, padx=15, pady=5)

            # Valor data
            vconcept = ws_act.cell(row=j, column=2).value
            concepte_label = tk.Label(window_grid, text=vconcept)
            concepte_label.grid(column=2, row=2, sticky=tk.E, padx=15, pady=5)

            # Valor classificació
            llista_concept = list(diccionari.keys())
            llista_items = tk.StringVar(value=llista_concept)
            l_res = tk.Listbox(window_grid, height=15, listvariable=llista_items, )
            l_res.grid(column=3, row=2, sticky=tk.E, padx=15, pady=5)
            l_res.bind('<<ListboxSelect>>', select_item)

            # Boto actualitza valor
            update_button = tk.Button(window_grid, text='Següent')
            update_button.grid(column=3, row=3, sticky=tk.E, padx=5, pady=5)
            update_button.bind('<ButtonRelease-1>', lambda e: window_grid.destroy())

            # window_grid.destroy()
            window_grid.mainloop()


# A partir d'una llista d'excels els agrupa en un excel
def combiexcel(llista):
    # Carreguem l'excel de comptes
    ex_comptes = load_workbook('/EstatComptes.xlsx')
    for document in llista:
        # Carreguem l'excel del banc
        ex_caixa = load_workbook(document)
        sheet_caixa = ex_caixa['in']

        # Creació fulla segons el mes
        data_mes = str(sheet_caixa['B4'].value)[0:10].split('-')[1]
        nom_fulla = Diccionari.mes.get(data_mes)
        if nom_fulla not in ex_comptes.sheetnames:
            ws1 = ex_comptes.create_sheet(nom_fulla)
            ws1.title = nom_fulla
            ws2 = ex_comptes.active = ex_comptes[nom_fulla]

            # calculate total number of rows and
            # columns in source Excel file
            mr = sheet_caixa.max_row
            mc = sheet_caixa.max_column

            # copying the cell values from source
            # Excel file to destination Excel file
            for i in range(3, mr + 1):
                for j in range(1, mc + 1):
                    if sheet_caixa.cell(row=i, column=j).value is not None:
                        # reading cell value from source excel file
                        c = sheet_caixa.cell(row=i, column=j)
                        # writing the read value to destination excel file
                        ws2.cell(row=i - 2, column=j).value = c.value

            for i in range(2, mr - 1):
                # Format data
                split_data = "/".join(list(reversed(str(ws2.cell(row=i, column=2).value)[0:10].split('-'))))
                ws2.cell(row=i, column=2).value = split_data
                # Format import
                if ws2.cell(row=i, column=3).value is not None:
                    imports = ".".join((str(ws2.cell(row=i, column=3).value).split('.')))
                    ws2.cell(row=i, column=3).value = float(imports)
                # Format saldo
                if ws2.cell(row=i, column=4).value is not None:
                    saldo = ".".join("".join(str(ws2.cell(row=i, column=4).value)[0:-3].split('.')).split(','))
                    ws2.cell(row=i, column=4).value = float(saldo)

            # format titols columnes
            a1 = ws2['A1']
            b1 = ws2['B1']
            c1 = ws2['C1']
            d1 = ws2['D1']
            e1 = ws2['E1']
            e1.value = "Classificació"

            a1.font = Font(bold=True, size=15)
            b1.font = Font(bold=True, size=15)
            c1.font = Font(bold=True, size=15)
            d1.font = Font(bold=True, size=15)
            e1.font = Font(bold=True, size=15)

            check_concept(Diccionari.classificació, ex_comptes)

            # filtres
            maxrow = ws2.max_row
            ws2.auto_filter.ref = f"A1:E{maxrow}"
            ws2.auto_filter.add_filter_column(5, ["Menja", "Compres", "Transport"])
            ws2.auto_filter.add_sort_condition(f"B2:B{maxrow}")

            #Taula resum
            sheet1 = ex_comptes['Gener']
            for i in range(2, 20):
                for j in range(7, 9):
                    ws2.cell(row=i, column=j).value = sheet1.cell(row=i, column=j).value
            # Cel·la estalvis
            ws2['H19'] = f'=D{maxrow}-D2'



        ex_comptes.save(filename='/EstatComptes.xlsx')
